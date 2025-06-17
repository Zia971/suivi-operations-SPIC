"""
Module utilitaires pour SPIC 2.0 - VERSION STREAMLIT
Fonctions de calcul, visualisation timeline colorée et gestion des alertes
Intégration timeline horizontale, frises chronologiques et alertes intelligentes
Optimisé pour Streamlit avec cache et performance
"""

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any, Tuple
import base64
from io import BytesIO
import json
import streamlit as st
import config

# ============================================================================
# CALCULS D'AVANCEMENT ET STATUTS
# ============================================================================

@st.cache_data(ttl=300)  # Cache 5 minutes
def calculate_progress(phases: List[Dict]) -> float:
    """Calcule le pourcentage d'avancement basé sur les phases validées"""
    
    if not phases:
        return 0.0
    
    total_phases = len(phases)
    phases_validees = sum(1 for phase in phases if phase.get('est_validee', False))
    
    return round((phases_validees / total_phases) * 100, 1)

def calculate_status_from_phases(phases: List[Dict], type_operation: str) -> str:
    """Calcule le statut automatique basé sur l'avancement des phases"""
    
    if not phases:
        return "🟡 À l'étude"
    
    # Vérifier s'il y a des blocages actifs
    for phase in phases:
        if phase.get('blocage_actif', False):
            return "🔴 Bloqué"
    
    # Calcul du pourcentage d'avancement
    pourcentage = calculate_progress(phases)
    
    # Utiliser la fonction de config.py
    return config.calculate_status_from_phases(phases, type_operation)

def get_current_phase(phases: List[Dict]) -> Dict:
    """Récupère la phase actuelle (première non validée)"""
    
    phases_triees = sorted(phases, key=lambda x: x.get('ordre_phase', 0))
    
    for phase in phases_triees:
        if not phase.get('est_validee', False):
            return phase
    
    # Si toutes les phases sont validées, retourner la dernière
    if phases_triees:
        return phases_triees[-1]
    
    return {}

def get_next_phases(phases: List[Dict], nb_phases: int = 3) -> List[Dict]:
    """Récupère les prochaines phases à traiter"""
    
    phases_triees = sorted(phases, key=lambda x: x.get('ordre_phase', 0))
    prochaines_phases = []
    
    phase_actuelle_trouvee = False
    
    for phase in phases_triees:
        if not phase.get('est_validee', False):
            if not phase_actuelle_trouvee:
                phase_actuelle_trouvee = True
                continue  # Skip la phase actuelle
            
            prochaines_phases.append(phase)
            if len(prochaines_phases) >= nb_phases:
                break
    
    return prochaines_phases

def detect_delays(phases: List[Dict]) -> List[Dict]:
    """Détecte les phases en retard"""
    
    phases_en_retard = []
    today = datetime.now().date()
    
    for phase in phases:
        if (not phase.get('est_validee', False) and 
            phase.get('date_fin_prevue')):
            
            try:
                date_fin = datetime.strptime(phase['date_fin_prevue'], '%Y-%m-%d').date()
                if today > date_fin:
                    jours_retard = (today - date_fin).days
                    phase_retard = phase.copy()
                    phase_retard['jours_retard'] = jours_retard
                    phases_en_retard.append(phase_retard)
            except:
                continue
    
    return sorted(phases_en_retard, key=lambda x: x.get('jours_retard', 0), reverse=True)

# ============================================================================
# SYSTÈME D'ALERTES ET CALCUL DE RISQUES
# ============================================================================

def check_alerts(operation_id: int, db_manager, include_suggestions: bool = True) -> List[Dict]:
    """Vérifie et génère les alertes pour une opération"""
    
    try:
        # Récupérer les données de l'opération
        operation = db_manager.get_operation_detail(operation_id)
        phases = db_manager.get_phases_by_operation(operation_id)
        
        if not operation or not phases:
            return []
        
        alertes = []
        today = datetime.now().date()
        
        # 1. Alertes de retard sur phases
        phases_retard = detect_delays(phases)
        for phase in phases_retard:
            alertes.append({
                'type_alerte': 'RETARD',
                'niveau_urgence': min(5, max(1, phase['jours_retard'] // 7 + 1)),
                'message': f"Phase '{phase['sous_phase']}' en retard de {phase['jours_retard']} jour(s)",
                'phase_concernee': phase['sous_phase'],
                'couleur': config.TYPES_ALERTES['RETARD']['couleur'],
                'icone': config.TYPES_ALERTES['RETARD']['icone']
            })
        
        # 2. Alertes de blocage actif
        phases_bloquees = [p for p in phases if p.get('blocage_actif', False)]
        for phase in phases_bloquees:
            alertes.append({
                'type_alerte': 'BLOCAGE',
                'niveau_urgence': 5,
                'message': f"Blocage actif sur '{phase['sous_phase']}' - {phase.get('motif_blocage', 'Motif non spécifié')}",
                'phase_concernee': phase['sous_phase'],
                'couleur': config.TYPES_ALERTES['BLOCAGE']['couleur'],
                'icone': config.TYPES_ALERTES['BLOCAGE']['icone']
            })
        
        # 3. Alertes d'échéance proche
        for phase in phases:
            if (not phase.get('est_validee', False) and 
                phase.get('date_fin_prevue')):
                
                try:
                    date_fin = datetime.strptime(phase['date_fin_prevue'], '%Y-%m-%d').date()
                    jours_restants = (date_fin - today).days
                    
                    if 0 <= jours_restants <= 7:
                        alertes.append({
                            'type_alerte': 'ATTENTION',
                            'niveau_urgence': 3,
                            'message': f"Échéance proche pour '{phase['sous_phase']}' dans {jours_restants} jour(s)",
                            'phase_concernee': phase['sous_phase'],
                            'couleur': config.TYPES_ALERTES['ATTENTION']['couleur'],
                            'icone': config.TYPES_ALERTES['ATTENTION']['icone']
                        })
                except:
                    continue
        
        # 4. Alertes de performance (suggestions)
        if include_suggestions:
            avancement = operation.get('pourcentage_avancement', 0)
            
            if avancement == 0 and operation.get('statut_principal') != '🟡 À l\'étude':
                alertes.append({
                    'type_alerte': 'ATTENTION',
                    'niveau_urgence': 2,
                    'message': "Aucune phase validée - Vérifiez l'avancement réel",
                    'phase_concernee': None,
                    'couleur': config.TYPES_ALERTES['ATTENTION']['couleur'],
                    'icone': '📊'
                })
            
            if avancement < 30 and '🚧 En travaux' in operation.get('statut_principal', ''):
                alertes.append({
                    'type_alerte': 'ATTENTION',
                    'niveau_urgence': 3,
                    'message': "Travaux peu avancés par rapport au statut déclaré",
                    'phase_concernee': None,
                    'couleur': config.TYPES_ALERTES['ATTENTION']['couleur'],
                    'icone': '🚧'
                })
        
        # Trier par niveau d'urgence décroissant
        alertes.sort(key=lambda x: x['niveau_urgence'], reverse=True)
        
        return alertes
        
    except Exception as e:
        st.error(f"❌ Erreur vérification alertes : {e}")
        return []

def calculate_risk_score(operation: Dict, phases: List[Dict], alertes: List[Dict] = None) -> float:
    """Calcule le score de risque d'une opération (utilise config.py)"""
    
    return config.calculate_risk_score(operation, phases, alertes)

def get_top_risk_operations(operations: List[Dict], db_manager, limit: int = 3) -> List[Dict]:
    """Identifie les opérations les plus à risque avec justifications détaillées"""
    
    operations_avec_risque = []
    
    for operation in operations:
        if operation.get('statut_principal') == '✅ Clôturé (soldé)':
            continue
        
        try:
            # Calculer le score de risque
            phases = db_manager.get_phases_by_operation(operation['id'])
            alertes = check_alerts(operation['id'], db_manager, include_suggestions=False)
            
            score_risque = calculate_risk_score(operation, phases, alertes)
            
            # Ajouter les justifications
            justifications = []
            
            if operation.get('has_active_blocage', False):
                justifications.append("🔴 Blocage actif")
            
            phases_retard = detect_delays(phases)
            if phases_retard:
                justifications.append(f"⏰ {len(phases_retard)} phase(s) en retard")
            
            if len(alertes) > 0:
                justifications.append(f"🚨 {len(alertes)} alerte(s)")
            
            avancement = operation.get('pourcentage_avancement', 0)
            if avancement == 0:
                justifications.append("📊 Aucun avancement")
            elif avancement < 30 and '🚧 En travaux' in operation.get('statut_principal', ''):
                justifications.append("🚧 Travaux peu avancés")
            
            if not justifications:
                justifications.append("📈 Risque calculé selon critères")
            
            operation_risk = operation.copy()
            operation_risk['score_risque'] = score_risque
            operation_risk['justifications'] = justifications
            operation_risk['nb_alertes'] = len(alertes)
            operation_risk['nb_phases_retard'] = len(phases_retard)
            
            operations_avec_risque.append(operation_risk)
            
        except Exception as e:
            st.error(f"⚠️ Erreur calcul risque pour {operation.get('nom', 'N/A')} : {e}")
            continue
    
    # Trier par score de risque décroissant
    operations_avec_risque.sort(key=lambda x: x['score_risque'], reverse=True)
    
    return operations_avec_risque[:limit]

# ============================================================================
# VISUALISATIONS TIMELINE ET FRISES CHRONOLOGIQUES
# ============================================================================

@st.cache_data(ttl=600)  # Cache 10 minutes
def generate_timeline(operation_id: int, timeline_data: Dict, type_viz: str = "chronologique") -> str:
    """Génère une timeline colorée et interactive pour Streamlit"""
    
    try:
        phases = timeline_data.get('phases', [])
        journal_entries = timeline_data.get('journal_entries', [])
        alertes = timeline_data.get('alertes', [])
        
        if not phases:
            return "<p>Aucune phase trouvée pour cette opération</p>"
        
        if type_viz == "chronologique":
            return _generate_chronological_timeline_streamlit(phases, journal_entries, alertes)
        elif type_viz == "mental_map":
            return _generate_mental_map_streamlit(phases, timeline_data.get('domaines', {}))
        else:
            return _generate_chronological_timeline_streamlit(phases, journal_entries, alertes)
            
    except Exception as e:
        st.error(f"❌ Erreur génération timeline : {e}")
        return f"<p>Erreur lors de la génération : {str(e)}</p>"

def _generate_chronological_timeline_streamlit(phases: List[Dict], journal_entries: List[Dict], alertes: List[Dict]) -> str:
    """Génère une frise chronologique horizontale colorée optimisée Streamlit"""
    
    try:
        # Préparer les données pour Plotly
        timeline_data = []
        today = datetime.now().date()
        
        # Ajouter les phases
        for phase in phases:
            # Déterminer les dates
            date_debut = None
            date_fin = None
            
            if phase.get('date_debut_prevue'):
                try:
                    date_debut = datetime.strptime(phase['date_debut_prevue'], '%Y-%m-%d').date()
                except:
                    pass
            
            if phase.get('date_fin_prevue'):
                try:
                    date_fin = datetime.strptime(phase['date_fin_prevue'], '%Y-%m-%d').date()
                except:
                    pass
            
            # Si pas de dates, estimer
            if not date_debut and not date_fin:
                duree = phase.get('duree_maxi_jours', 30)
                if timeline_data:
                    derniere_fin = timeline_data[-1].get('Finish', today)
                    if isinstance(derniere_fin, str):
                        derniere_fin = datetime.strptime(derniere_fin, '%Y-%m-%d').date()
                    date_debut = derniere_fin + timedelta(days=1)
                else:
                    date_debut = today
                date_fin = date_debut + timedelta(days=duree)
            elif date_debut and not date_fin:
                duree = phase.get('duree_maxi_jours', 30)
                date_fin = date_debut + timedelta(days=duree)
            elif not date_debut and date_fin:
                duree = phase.get('duree_maxi_jours', 30)
                date_debut = date_fin - timedelta(days=duree)
            
            # Déterminer la couleur selon l'état
            if phase.get('blocage_actif', False):
                couleur = '#dc2626'  # Rouge - Bloqué
                statut_text = "🔴 BLOQUÉ"
            elif phase.get('est_validee', False):
                couleur = '#10b981'  # Vert - Validé
                statut_text = "✅ VALIDÉ"
            elif date_fin and today > date_fin and not phase.get('est_validee', False):
                couleur = '#ef4444'  # Rouge - En retard
                statut_text = "⏰ EN RETARD"
            elif date_fin and (date_fin - today).days <= 7 and not phase.get('est_validee', False):
                couleur = '#f59e0b'  # Orange - Échéance proche
                statut_text = "⚡ URGENT"
            else:
                couleur = phase.get('couleur_domaine', '#6b7280')  # Couleur selon domaine
                statut_text = "⏳ EN COURS"
            
            # Icône selon le domaine
            domaine = phase.get('domaine', 'OPERATIONNEL')
            icone_domaine = config.DOMAINES_OPERATIONNELS.get(domaine, {}).get('icone', '📌')
            
            timeline_data.append({
                'Task': f"{icone_domaine} {phase.get('sous_phase', 'Phase')}",
                'Start': date_debut,
                'Finish': date_fin,
                'Resource': phase.get('responsable_principal', 'Non défini'),
                'Statut': statut_text,
                'Domaine': domaine,
                'Description': f"{phase.get('phase_principale', '')} - {phase.get('responsable_principal', '')}",
                'Couleur': couleur,
                'Type': 'Phase'
            })
        
        # Ajouter les événements du journal importants
        for entry in journal_entries[:5]:  # Top 5 entrées importantes
            try:
                date_entry = datetime.strptime(entry['date_saisie'], '%Y-%m-%d %H:%M:%S').date()
                
                timeline_data.append({
                    'Task': f"📝 {entry.get('type_action', 'INFO')}: {entry.get('contenu', '')[:30]}...",
                    'Start': date_entry,
                    'Finish': date_entry + timedelta(days=1),  # Durée d'1 jour pour visibilité
                    'Resource': entry.get('auteur', 'N/A'),
                    'Statut': 'JOURNAL',
                    'Domaine': 'Journal',
                    'Description': entry.get('contenu', ''),
                    'Couleur': entry.get('couleur_timeline', '#6b7280'),
                    'Type': 'Journal'
                })
            except:
                continue
        
        # Ajouter les alertes actives
        for alerte in alertes[:3]:  # Top 3 alertes
            try:
                date_alerte = datetime.strptime(alerte['date_creation'], '%Y-%m-%d %H:%M:%S').date()
                
                timeline_data.append({
                    'Task': f"🚨 {alerte.get('type_alerte', 'ALERTE')}: {alerte.get('message', '')[:25]}...",
                    'Start': date_alerte,
                    'Finish': date_alerte + timedelta(days=2),  # Durée de 2 jours pour visibilité
                    'Resource': 'Système',
                    'Statut': 'ALERTE',
                    'Domaine': 'Alertes',
                    'Description': alerte.get('message', ''),
                    'Couleur': alerte.get('couleur_alerte', '#ef4444'),
                    'Type': 'Alerte'
                })
            except:
                continue
        
        # Créer le graphique Plotly
        if not timeline_data:
            return "<p>Impossible de générer la timeline</p>"
        
        df = pd.DataFrame(timeline_data)
        
        # Couleurs personnalisées selon le type
        color_map = {}
        for item in timeline_data:
            color_map[item['Statut']] = item['Couleur']
        
        fig = px.timeline(
            df,
            x_start="Start",
            x_end="Finish",
            y="Task",
            color="Statut",
            title="📅 Timeline Enrichie - Phases, Journal & Alertes",
            hover_data=["Resource", "Description", "Domaine"],
            color_discrete_map=color_map,
            height=max(600, len(timeline_data) * 35)
        )
        
        # Personnaliser l'affichage
        fig.update_layout(
            xaxis_title="Période",
            yaxis_title="Éléments Timeline",
            font=dict(size=11),
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            ),
            plot_bgcolor='white',
            paper_bgcolor='white'
        )
        
        fig.update_yaxes(autorange="reversed")
        
        # Ajouter une ligne verticale pour aujourd'hui
        fig.add_vline(
            x=today,
            line_dash="dash",
            line_color="red",
            line_width=2,
            annotation_text="Aujourd'hui",
            annotation_position="top"
        )
        
        # Ajouter annotations pour les domaines
        domaines_y_positions = {}
        for i, item in enumerate(timeline_data):
            if item['Type'] == 'Phase':
                domaine = item['Domaine']
                if domaine not in domaines_y_positions:
                    domaines_y_positions[domaine] = []
                domaines_y_positions[domaine].append(i)
        
        return fig.to_html(include_plotlyjs='cdn', div_id=f"timeline_{datetime.now().timestamp()}")
        
    except Exception as e:
        st.error(f"❌ Erreur timeline chronologique : {e}")
        return f"<p>Erreur génération timeline : {str(e)}</p>"

def _generate_mental_map_streamlit(phases: List[Dict], domaines: Dict) -> str:
    """Génère une carte mentale interactive avec les phases groupées par domaines"""
    
    try:
        # Regrouper par domaine
        phases_par_domaine = {}
        for phase in phases:
            domaine = phase.get('domaine', 'OPERATIONNEL')
            if domaine not in phases_par_domaine:
                phases_par_domaine[domaine] = []
            phases_par_domaine[domaine].append(phase)
        
        # Générer HTML de carte mentale moderne
        html_content = f"""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 25px; border-radius: 15px; color: white; font-family: Arial, sans-serif;">
            <h3 style="text-align: center; margin-bottom: 25px; font-size: 1.5em;">🧠 Carte Mentale par Domaines</h3>
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px;">
        """
        
        for domaine, sous_phases in phases_par_domaine.items():
            phases_validees = sum(1 for p in sous_phases if p.get('est_validee', False))
            total_phases = len(sous_phases)
            pourcentage = (phases_validees / total_phases) * 100 if total_phases > 0 else 0
            
            # Informations du domaine depuis config
            domaine_info = domaines.get(domaine, {})
            couleur_domaine = domaine_info.get('couleur', '#6b7280')
            icone_domaine = domaine_info.get('icone', '📌')
            nom_domaine = domaine_info.get('nom', domaine)
            
            # Couleur selon avancement
            if pourcentage == 100:
                couleur_fond = "#10b981"  # Vert
                status_icone = "✅"
            elif pourcentage >= 50:
                couleur_fond = "#f59e0b"  # Orange
                status_icone = "⚡"
            elif pourcentage > 0:
                couleur_fond = "#3b82f6"  # Bleu
                status_icone = "🔄"
            else:
                couleur_fond = "#6b7280"  # Gris
                status_icone = "⏳"
            
            html_content += f"""
            <div style="background: rgba(255,255,255,0.15); padding: 20px; border-radius: 12px; border-left: 5px solid {couleur_domaine}; backdrop-filter: blur(10px);">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
                    <h4 style="margin: 0; color: white; display: flex; align-items: center; gap: 8px;">
                        {icone_domaine} {nom_domaine}
                    </h4>
                    <div style="display: flex; align-items: center; gap: 10px;">
                        <span style="background: {couleur_fond}; padding: 6px 12px; border-radius: 20px; font-weight: bold; font-size: 0.9em;">
                            {phases_validees}/{total_phases}
                        </span>
                        <span style="font-size: 1.2em;">{status_icone}</span>
                    </div>
                </div>
                
                <div style="margin-bottom: 10px;">
                    <div style="background: rgba(255,255,255,0.2); border-radius: 10px; overflow: hidden; height: 8px;">
                        <div style="background: {couleur_fond}; height: 100%; width: {pourcentage}%; transition: width 0.3s ease;"></div>
                    </div>
                    <span style="font-size: 0.8em; color: rgba(255,255,255,0.8);">{pourcentage:.0f}% complété</span>
                </div>
                
                <div style="display: grid; gap: 8px; margin-left: 10px; max-height: 200px; overflow-y: auto;">
            """
            
            for sous_phase in sous_phases:
                if sous_phase.get('blocage_actif', False):
                    statut_icone = "🔴"
                    statut_couleur = "#ef4444"
                    statut_text = "Bloqué"
                elif sous_phase.get('est_validee', False):
                    statut_icone = "✅"
                    statut_couleur = "#10b981"
                    statut_text = "Validé"
                else:
                    # Vérifier si en retard
                    statut_icone = "⏳"
                    statut_couleur = "#6b7280"
                    statut_text = "En cours"
                    
                    if sous_phase.get('date_fin_prevue'):
                        try:
                            date_fin = datetime.strptime(sous_phase['date_fin_prevue'], '%Y-%m-%d').date()
                            if datetime.now().date() > date_fin:
                                statut_icone = "⏰"
                                statut_couleur = "#ef4444"
                                statut_text = "En retard"
                        except:
                            pass
                
                html_content += f"""
                <div style="display: flex; align-items: center; gap: 12px; padding: 8px; background: rgba(255,255,255,0.1); border-radius: 8px; transition: background 0.2s ease;">
                    <span style="color: {statut_couleur}; font-size: 16px; min-width: 20px;">{statut_icone}</span>
                    <div style="flex: 1;">
                        <span style="color: rgba(255,255,255,0.95); font-size: 0.9em;">{sous_phase.get('sous_phase', 'Phase')}</span>
                        <div style="color: rgba(255,255,255,0.7); font-size: 0.7em;">{statut_text}</div>
                    </div>
                    {f'<span style="color: rgba(255,255,255,0.6); font-size: 0.7em;">💰</span>' if sous_phase.get('impact_rem', False) else ''}
                </div>
                """
            
            html_content += "</div></div>"
        
        html_content += """
            </div>
            <div style="margin-top: 20px; text-align: center; color: rgba(255,255,255,0.8); font-size: 0.8em;">
                💰 = Impact sur la REM | Mise à jour temps réel
            </div>
        </div>
        """
        
        return html_content
        
    except Exception as e:
        st.error(f"❌ Erreur carte mentale : {e}")
        return f"<p>Erreur génération carte mentale : {str(e)}</p>"

# ============================================================================
# GRAPHIQUES ET KPI POUR STREAMLIT
# ============================================================================

@st.cache_data(ttl=300)
def generate_kpi_charts_streamlit(kpi_data: Dict) -> Dict[str, go.Figure]:
    """Génère les graphiques KPI optimisés pour Streamlit"""
    
    charts = {}
    
    try:
        # 1. Graphique répartition statuts (Donut moderne)
        if 'repartition_statuts' in kpi_data and kpi_data['repartition_statuts']:
            statuts = list(kpi_data['repartition_statuts'].keys())
            valeurs = list(kpi_data['repartition_statuts'].values())
            
            # Couleurs selon les statuts
            couleurs = []
            for statut in statuts:
                couleurs.append(config.STATUTS_GLOBAUX.get(statut, {}).get('couleur', '#6b7280'))
            
            fig_statuts = go.Figure(data=[go.Pie(
                labels=statuts,
                values=valeurs,
                hole=.4,
                marker_colors=couleurs,
                textinfo='label+percent',
                textposition='outside',
                hovertemplate='<b>%{label}</b><br>' +
                              'Nombre: %{value}<br>' +
                              'Pourcentage: %{percent}<br>' +
                              '<extra></extra>'
            )])
            
            fig_statuts.update_layout(
                title={
                    'text': "📊 Répartition par Statut",
                    'x': 0.5,
                    'xanchor': 'center'
                },
                font=dict(size=12),
                height=400,
                margin=dict(t=50, b=50, l=50, r=50)
            )
            
            charts['statuts'] = fig_statuts
        
         # 2. Graphique répartition types (Barres colorées)
        if 'repartition_types' in kpi_data and kpi_data['repartition_types']:
            types = list(kpi_data['repartition_types'].keys())
            valeurs = list(kpi_data['repartition_types'].values())
            couleurs_types = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6']
            
            fig_types = go.Figure(data=[go.Bar(
                x=types,
                y=valeurs,
                marker_color=couleurs_types[:len(types)],
                text=valeurs,
                textposition='auto',
                hovertemplate='<b>%{x}</b><br>' +
                              'Nombre: %{y}<br>' +
                              '<extra></extra>'
            )])
            
            fig_types.update_layout(
                title={
                    'text': "📈 Opérations par Type",
                    'x': 0.5,
                    'xanchor': 'center'
                },
                xaxis_title="Type d'opération",
                yaxis_title="Nombre",
                font=dict(size=12),
                height=400,
                margin=dict(t=50, b=50, l=50, r=50),
                plot_bgcolor='white'
            )
            
            charts['types'] = fig_types
        
        # 3. Graphique avancement global (Jauge moderne)
        if 'avancement_moyen' in kpi_data:
            avancement = kpi_data['avancement_moyen']
            
            fig_avancement = go.Figure(go.Indicator(
                mode = "gauge+number+delta",
                value = avancement,
                domain = {'x': [0, 1], 'y': [0, 1]},
                title = {'text': "📊 Avancement Moyen Global", 'font': {'size': 16}},
                delta = {'reference': 50, 'position': "top"},
                gauge = {
                    'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
                    'bar': {'color': "darkblue"},
                    'bgcolor': "white",
                    'borderwidth': 2,
                    'bordercolor': "gray",
                    'steps': [
                        {'range': [0, 25], 'color': '#ef4444'},
                        {'range': [25, 50], 'color': '#f59e0b'},
                        {'range': [50, 75], 'color': '#3b82f6'},
                        {'range': [75, 100], 'color': '#10b981'}
                    ],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': 90
                    }
                }
            ))
            
            fig_avancement.update_layout(
                height=400,
                margin=dict(t=50, b=50, l=50, r=50)
            )
            
            charts['avancement'] = fig_avancement
        
        # 4. Graphique évolution temporelle (si données disponibles)
        if all(key in kpi_data for key in ['nouvelles_operations_semaine', 'operations_cloturees_semaine']):
            categories = ['Nouvelles', 'Clôturées', 'En retard']
            valeurs = [
                kpi_data['nouvelles_operations_semaine'],
                kpi_data['operations_cloturees_semaine'],
                kpi_data.get('operations_en_retard', 0)
            ]
            couleurs = ['#10b981', '#3b82f6', '#ef4444']
            
            fig_evolution = go.Figure(data=[go.Bar(
                x=categories,
                y=valeurs,
                marker_color=couleurs,
                text=valeurs,
                textposition='auto',
                hovertemplate='<b>%{x}</b><br>' +
                              'Nombre: %{y}<br>' +
                              '<extra></extra>'
            )])
            
            fig_evolution.update_layout(
                title={
                    'text': "📈 Évolution Dernière Semaine",
                    'x': 0.5,
                    'xanchor': 'center'
                },
                xaxis_title="Catégorie",
                yaxis_title="Nombre",
                font=dict(size=12),
                height=400,
                margin=dict(t=50, b=50, l=50, r=50),
                plot_bgcolor='white'
            )
            
            charts['evolution'] = fig_evolution
        
        # 5. Graphique REM si disponible
        if 'rem_totale_portfolio' in kpi_data and kpi_data['rem_totale_portfolio'] > 0:
            rem_totale = kpi_data['rem_totale_portfolio']
            ops_avec_rem = kpi_data.get('operations_avec_rem', 0)
            
            fig_rem = go.Figure(go.Indicator(
                mode = "number",
                value = rem_totale,
                title = {"text": f"💰 REM Portfolio<br><span style='font-size:12px'>{ops_avec_rem} opérations</span>"},
                number = {'suffix': " €", 'font': {'size': 24}},
                domain = {'x': [0, 1], 'y': [0, 1]}
            ))
            
            fig_rem.update_layout(
                height=200,
                margin=dict(t=50, b=50, l=50, r=50)
            )
            
            charts['rem'] = fig_rem
        
        return charts
        
    except Exception as e:
        st.error(f"❌ Erreur génération graphiques KPI : {e}")
        return {}

def create_risk_analysis_chart_streamlit(operations_risk: List[Dict]) -> go.Figure:
    """Crée un graphique d'analyse des risques optimisé Streamlit"""
    
    try:
        if not operations_risk:
            fig = go.Figure()
            fig.add_annotation(
                text="Aucune donnée de risque disponible",
                xref="paper", yref="paper",
                x=0.5, y=0.5, xanchor='center', yanchor='middle',
                showarrow=False, font=dict(size=16)
            )
            return fig
        
        # Préparer les données
        noms = [op.get('nom', 'N/A')[:25] + '...' if len(op.get('nom', '')) > 25 else op.get('nom', 'N/A') for op in operations_risk]
        scores = [op.get('score_risque', 0) for op in operations_risk]
        justifications = [', '.join(op.get('justifications', [])) for op in operations_risk]
        
        # Couleurs selon le niveau de risque
        couleurs = []
        for score in scores:
            if score >= 80:
                couleurs.append('#dc2626')  # Rouge - Risque très élevé
            elif score >= 60:
                couleurs.append('#f59e0b')  # Orange - Risque élevé
            elif score >= 40:
                couleurs.append('#3b82f6')  # Bleu - Risque modéré
            else:
                couleurs.append('#10b981')  # Vert - Risque faible
        
        fig = go.Figure(data=[go.Bar(
            x=noms,
            y=scores,
            marker_color=couleurs,
            text=[f"{score:.0f}" for score in scores],
            textposition='auto',
            hovertemplate='<b>%{x}</b><br>' +
                          'Score de Risque: %{y:.1f}<br>' +
                          'Justifications: %{customdata}<br>' +
                          '<extra></extra>',
            customdata=justifications
        )])
        
        fig.update_layout(
            title={
                'text': "🚨 Analyse des Risques par Opération",
                'x': 0.5,
                'xanchor': 'center'
            },
            xaxis_title="Opérations",
            yaxis_title="Score de Risque (0-100)",
            xaxis={'tickangle': -45},
            height=450,
            margin=dict(t=60, b=100, l=60, r=60),
            plot_bgcolor='white'
        )
        
        # Ligne de seuil critique
        fig.add_hline(y=70, line_dash="dash", line_color="red", 
                      annotation_text="Seuil critique", annotation_position="bottom right")
        
        return fig
        
    except Exception as e:
        st.error(f"❌ Erreur graphique analyse risques : {e}")
        return go.Figure()

def create_progress_distribution_chart_streamlit(operations: List[Dict]) -> go.Figure:
    """Crée un histogramme de distribution des avancements pour Streamlit"""
    
    try:
        if not operations:
            fig = go.Figure()
            fig.add_annotation(
                text="Aucune donnée d'avancement disponible",
                xref="paper", yref="paper",
                x=0.5, y=0.5, xanchor='center', yanchor='middle',
                showarrow=False, font=dict(size=16)
            )
            return fig
        
        # Préparer les données d'avancement
        avancements = [op.get('pourcentage_avancement', 0) for op in operations]
        
        # Créer des bins pour l'histogramme
        bins = [0, 20, 40, 60, 80, 100]
        labels = ['0-20%', '20-40%', '40-60%', '60-80%', '80-100%']
        couleurs = ['#ef4444', '#f59e0b', '#eab308', '#3b82f6', '#10b981']
        
        # Compter les opérations par tranche
        counts = []
        for i in range(len(bins)-1):
            if i == len(bins)-2:  # Dernière tranche inclut 100%
                count = sum(1 for av in avancements if bins[i] <= av <= bins[i+1])
            else:
                count = sum(1 for av in avancements if bins[i] <= av < bins[i+1])
            counts.append(count)
        
        fig = go.Figure(data=[go.Bar(
            x=labels,
            y=counts,
            marker_color=couleurs,
            text=counts,
            textposition='auto',
            hovertemplate='<b>%{x}</b><br>' +
                          'Nombre d\'opérations: %{y}<br>' +
                          '<extra></extra>'
        )])
        
        fig.update_layout(
            title={
                'text': "📊 Distribution des Avancements",
                'x': 0.5,
                'xanchor': 'center'
            },
            xaxis_title="Avancement (%)",
            yaxis_title="Nombre d'opérations",
            height=400,
            margin=dict(t=50, b=50, l=50, r=50),
            plot_bgcolor='white'
        )
        
        return fig
        
    except Exception as e:
        st.error(f"❌ Erreur graphique distribution : {e}")
        return go.Figure()

def create_rem_evolution_chart(rem_data: List[Dict]) -> go.Figure:
    """Crée un graphique d'évolution REM pour Streamlit"""
    
    try:
        if not rem_data:
            fig = go.Figure()
            fig.add_annotation(
                text="Aucune donnée REM disponible",
                xref="paper", yref="paper",
                x=0.5, y=0.5, xanchor='center', yanchor='middle',
                showarrow=False, font=dict(size=16)
            )
            return fig
        
        # Préparer les données
        dates = [item.get('date_calcul', '') for item in rem_data]
        rem_values = [item.get('rem_annuelle_prevue', 0) for item in rem_data]
        operations = [item.get('operation_nom', 'N/A') for item in rem_data]
        
        fig = go.Figure()
        
        # Ligne principale
        fig.add_trace(go.Scatter(
            x=dates,
            y=rem_values,
            mode='lines+markers',
            name='REM Annuelle Prévue',
            line=dict(color='#10b981', width=3),
            marker=dict(size=8, color='#10b981'),
            hovertemplate='<b>%{customdata}</b><br>' +
                          'Date: %{x}<br>' +
                          'REM: %{y:,.0f} €<br>' +
                          '<extra></extra>',
            customdata=operations
        ))
        
        fig.update_layout(
            title={
                'text': "💰 Évolution des Projections REM",
                'x': 0.5,
                'xanchor': 'center'
            },
            xaxis_title="Période",
            yaxis_title="REM Annuelle (€)",
            height=400,
            margin=dict(t=50, b=50, l=50, r=50),
            plot_bgcolor='white'
        )
        
        return fig
        
    except Exception as e:
        st.error(f"❌ Erreur graphique REM : {e}")
        return go.Figure()

# ============================================================================
# FORMATAGE ET UTILITAIRES STREAMLIT
# ============================================================================

def format_currency(montant: float, devise: str = "€") -> str:
    """Formate un montant en devise avec séparateurs français"""
    
    try:
        if montant == 0:
            return f"0 {devise}"
        
        # Formatage français avec espaces comme séparateurs de milliers
        if abs(montant) >= 1:
            montant_formate = f"{montant:,.0f}".replace(",", " ")
        else:
            montant_formate = f"{montant:.2f}".replace(".", ",")
        
        return f"{montant_formate} {devise}"
        
    except:
        return f"{montant} {devise}"

def format_date_fr(date_input, format_sortie: str = "court") -> str:
    """Formate une date en français DD/MM/YYYY"""
    
    if not date_input:
        return "Date non définie"
    
    try:
        if isinstance(date_input, str):
            if len(date_input) == 10 and date_input.count('-') == 2:
                # Format YYYY-MM-DD
                date_obj = datetime.strptime(date_input, '%Y-%m-%d').date()
            elif len(date_input) > 10:
                # Format avec heure
                date_obj = datetime.strptime(date_input[:10], '%Y-%m-%d').date()
            else:
                return date_input
        elif hasattr(date_input, 'strftime'):
            date_obj = date_input
        else:
            return str(date_input)
        
        if format_sortie == "long":
            mois_fr = [
                "", "janvier", "février", "mars", "avril", "mai", "juin",
                "juillet", "août", "septembre", "octobre", "novembre", "décembre"
            ]
            return f"{date_obj.day} {mois_fr[date_obj.month]} {date_obj.year}"
        else:
            return date_obj.strftime('%d/%m/%Y')
            
    except Exception as e:
        st.error(f"⚠️ Erreur formatage date {date_input} : {e}")
        return str(date_input)

def calculate_duration_working_days(date_debut: str, date_fin: str) -> int:
    """Calcule la durée en jours ouvrés entre deux dates"""
    
    try:
        if not date_debut or not date_fin:
            return 0
        
        debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        if fin < debut:
            return 0
        
        # Calcul simple sans jours fériés (peut être amélioré)
        total_days = (fin - debut).days
        
        # Approximation : retirer ~22% pour weekends (5/7 jours ouvrés)
        working_days = int(total_days * 5/7)
        
        return working_days
        
    except Exception as e:
        st.error(f"⚠️ Erreur calcul durée : {e}")
        return 0

def estimate_completion_date(phases: List[Dict], current_date: date = None) -> Optional[date]:
    """Estime la date de fin d'opération basée sur les phases restantes"""
    
    try:
        if not phases:
            return None
        
        if current_date is None:
            current_date = datetime.now().date()
        
        # Trouver la dernière phase non validée
        phases_restantes = [p for p in phases if not p.get('est_validee', False)]
        
        if not phases_restantes:
            # Toutes les phases sont validées
            return current_date
        
        # Prendre la date de fin prévue la plus tardive
        dates_fin = []
        
        for phase in phases_restantes:
            if phase.get('date_fin_prevue'):
                try:
                    date_fin = datetime.strptime(phase['date_fin_prevue'], '%Y-%m-%d').date()
                    dates_fin.append(date_fin)
                except:
                    continue
        
        if dates_fin:
            return max(dates_fin)
        
        # Si pas de dates, estimer basé sur les durées
        duree_restante = sum(phase.get('duree_maxi_jours', 30) for phase in phases_restantes)
        return current_date + timedelta(days=duree_restante)
        
    except Exception as e:
        st.error(f"⚠️ Erreur estimation date fin : {e}")
        return None

def get_phase_icon(phase: Dict) -> str:
    """Retourne l'icône appropriée pour une phase"""
    
    try:
        # Icônes selon l'état
        if phase.get('blocage_actif', False):
            return "🔴"
        elif phase.get('est_validee', False):
            return "✅"
        
        # Icônes selon le domaine
        domaine = phase.get('domaine', 'OPERATIONNEL')
        if domaine in config.DOMAINES_OPERATIONNELS:
            return config.DOMAINES_OPERATIONNELS[domaine]['icone']
        
        # Icônes selon le type de phase (fallback)
        phase_nom = phase.get('sous_phase', '').lower()
        
        if any(word in phase_nom for word in ['montage', 'opportunité']):
            return "🎯"
        elif any(word in phase_nom for word in ['études', 'esq', 'aps', 'apd', 'pro']):
            return "📐"
        elif any(word in phase_nom for word in ['autorisation', 'pc', 'permis']):
            return "📋"
        elif any(word in phase_nom for word in ['financement', 'prêt', 'cdc']):
            return "💰"
        elif any(word in phase_nom for word in ['consultation', 'cao', 'offres']):
            return "📢"
        elif any(word in phase_nom for word in ['attribution', 'marché', 'signature']):
            return "✍️"
        elif any(word in phase_nom for word in ['travaux', 'chantier', 'construction']):
            return "🚧"
        elif any(word in phase_nom for word in ['réception', 'livraison']):
            return "🎁"
        elif any(word in phase_nom for word in ['gpa', 'garantie']):
            return "🛡️"
        elif any(word in phase_nom for word in ['clôture', 'solde', 'archivage']):
            return "📁"
        else:
            return "📌"
            
    except Exception as e:
        st.error(f"⚠️ Erreur icône phase : {e}")
        return "📌"

# ============================================================================
# FONCTIONS SPÉCIALES POUR STREAMLIT
# ============================================================================

@st.cache_data(ttl=600)
def generate_operation_summary_cached(operation: Dict, phases: List[Dict], alertes: List[Dict] = None) -> Dict:
    """Génère un résumé complet d'une opération (version cachée pour Streamlit)"""
    
    try:
        summary = {
            'nom': operation.get('nom', 'N/A'),
            'type_operation': operation.get('type_operation', 'N/A'),
            'statut_principal': operation.get('statut_principal', 'N/A'),
            'responsable_aco': operation.get('responsable_aco', 'N/A'),
            'avancement': operation.get('pourcentage_avancement', 0),
        }
        
        # Analyse des phases
        if phases:
            total_phases = len(phases)
            phases_validees = sum(1 for p in phases if p.get('est_validee', False))
            phases_bloquees = sum(1 for p in phases if p.get('blocage_actif', False))
            phases_retard = len(detect_delays(phases))
            
            summary.update({
                'total_phases': total_phases,
                'phases_validees': phases_validees,
                'phases_bloquees': phases_bloquees,
                'phases_en_retard': phases_retard,
                'phase_actuelle': get_current_phase(phases).get('sous_phase', 'N/A'),
                'prochaines_phases': [p.get('sous_phase', 'N/A') for p in get_next_phases(phases, 2)]
            })
            
            # Estimation date de fin
            date_fin_estimee = estimate_completion_date(phases)
            if date_fin_estimee:
                summary['date_fin_estimee'] = format_date_fr(date_fin_estimee.isoformat())
        
        # Analyse des alertes
        if alertes:
            alertes_par_niveau = {}
            for alerte in alertes:
                niveau = alerte.get('niveau_urgence', 1)
                alertes_par_niveau[niveau] = alertes_par_niveau.get(niveau, 0) + 1
            
            summary.update({
                'total_alertes': len(alertes),
                'alertes_par_niveau': alertes_par_niveau,
                'alerte_max_urgence': max([a.get('niveau_urgence', 1) for a in alertes]) if alertes else 0
            })
        
        # Score de risque
        summary['score_risque'] = calculate_risk_score(operation, phases, alertes)
        
        # REM si disponible
        if operation.get('rem_annuelle_prevue', 0) > 0:
            summary['rem_annuelle'] = format_currency(operation['rem_annuelle_prevue'])
        
        return summary
        
    except Exception as e:
        st.error(f"❌ Erreur génération résumé : {e}")
        return {'erreur': str(e)}

def get_weekly_focus_tasks_streamlit(operations: List[Dict], db_manager) -> List[Dict]:
    """Récupère les tâches prioritaires de la semaine pour Streamlit"""
    
    try:
        taches_prioritaires = []
        today = datetime.now().date()
        fin_semaine = today + timedelta(days=7)
        
        for operation in operations:
            # Récupérer les phases de l'opération
            phases = db_manager.get_phases_by_operation(operation['id'])
            
            # Phases à échéance cette semaine
            for phase in phases:
                if (not phase.get('est_validee', False) and 
                    phase.get('date_fin_prevue')):
                    
                    try:
                        date_fin = datetime.strptime(phase['date_fin_prevue'], '%Y-%m-%d').date()
                        
                        if today <= date_fin <= fin_semaine:
                            taches_prioritaires.append({
                                'operation_nom': operation.get('nom', 'N/A'),
                                'operation_id': operation.get('id'),
                                'phase_nom': phase.get('sous_phase', 'N/A'),
                                'date_echeance': date_fin,
                                'jours_restants': (date_fin - today).days,
                                'priorite': 'URGENT' if (date_fin - today).days <= 2 else 'IMPORTANT',
                                'responsable': phase.get('responsable_principal', 'Non défini'),
                                'icone': get_phase_icon(phase),
                                'domaine': phase.get('domaine', 'OPERATIONNEL')
                            })
                    except:
                        continue
            
            # Phases bloquées (priorité absolue)
            phases_bloquees = [p for p in phases if p.get('blocage_actif', False)]
            for phase in phases_bloquees:
                taches_prioritaires.append({
                    'operation_nom': operation.get('nom', 'N/A'),
                    'operation_id': operation.get('id'),
                    'phase_nom': phase.get('sous_phase', 'N/A'),
                    'date_echeance': today,
                    'jours_restants': 0,
                    'priorite': 'BLOQUÉ',
                    'responsable': phase.get('responsable_principal', 'Non défini'),
                    'icone': '🔴',
                    'motif_blocage': phase.get('motif_blocage', 'Non spécifié'),
                    'domaine': phase.get('domaine', 'OPERATIONNEL')
                })
        
        # Trier par priorité puis par échéance
        ordre_priorite = {'BLOQUÉ': 0, 'URGENT': 1, 'IMPORTANT': 2}
        taches_prioritaires.sort(key=lambda x: (ordre_priorite.get(x['priorite'], 3), x['jours_restants']))
        
        return taches_prioritaires[:10]  # Top 10
        
    except Exception as e:
        st.error(f"❌ Erreur récupération tâches semaine : {e}")
        return []

@st.cache_data(ttl=300)
def generate_aco_performance_summary_streamlit(aco_nom: str, operations: List[Dict], performance_data: Dict = None) -> Dict:
    """Génère un résumé de performance pour un ACO optimisé Streamlit"""
    
    try:
        if not operations:
            return {'erreur': 'Aucune opération trouvée'}
        
        # Statistiques de base
        total_operations = len(operations)
        operations_actives = len([op for op in operations if op.get('statut_principal') != '✅ Clôturé (soldé)'])
        
        # Avancement moyen
        avancements = [op.get('pourcentage_avancement', 0) for op in operations]
        avancement_moyen = sum(avancements) / len(avancements) if avancements else 0
        
        # Répartition par statut
        statuts = {}
        for op in operations:
            statut = op.get('statut_principal', 'Non défini')
            statuts[statut] = statuts.get(statut, 0) + 1
        
        # Opérations à risque
        operations_a_risque = []
        total_alertes = 0
        
        for operation in operations:
            if operation.get('statut_principal') == '✅ Clôturé (soldé)':
                continue
                
            score_risque = operation.get('score_risque', 0)
            
            if (operation.get('has_active_blocage', False) or 
                score_risque > 50):
                operations_a_risque.append(operation.get('nom', 'N/A'))
        
        # REM totale gérée
        rem_totale = sum(op.get('rem_annuelle_prevue', 0) for op in operations)
        
        # Performance relative (si données disponibles)
        performance_relative = 0
        if performance_data:
            performance_relative = performance_data.get('performance_relative', 0)
        
        # Tendance (simulation basée sur les données)
        operations_en_avancement = len([op for op in operations if op.get('pourcentage_avancement', 0) > 50])
        tendance = "positive" if operations_en_avancement > total_operations * 0.6 else "neutre"
        
        if operations_a_risque:
            tendance = "attention"
        
        return {
            'aco_nom': aco_nom,
            'total_operations': total_operations,
            'operations_actives': operations_actives,
            'avancement_moyen': round(avancement_moyen, 1),
            'repartition_statuts': statuts,
            'operations_a_risque': operations_a_risque[:5],  # Top 5
            'nb_operations_a_risque': len(operations_a_risque),
            'rem_totale_geree': format_currency(rem_totale),
            'performance_relative': performance_relative,
            'tendance': tendance,
            'date_maj': datetime.now().strftime('%d/%m/%Y %H:%M')
        }
        
    except Exception as e:
        st.error(f"❌ Erreur résumé performance ACO : {e}")
        return {'erreur': str(e)}

# ============================================================================
# EXPORT ET UTILITAIRES AVANCÉS
# ============================================================================

def export_to_excel_streamlit(operations: List[Dict], phases_data: Dict = None) -> BytesIO:
    """Exporte les données vers Excel optimisé pour Streamlit (retourne BytesIO)"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Feuille 1 : Vue d'ensemble des opérations
        ws1 = wb.active
        ws1.title = "Operations"
        
        if operations:
            df_operations = pd.DataFrame(operations)
            
            # Colonnes à inclure
            colonnes_export = [
                'nom', 'type_operation', 'statut_principal', 'responsable_aco',
                'commune', 'pourcentage_avancement', 'date_creation', 'score_risque',
                'rem_annuelle_prevue', 'budget_total', 'nb_logements_total'
            ]
            
            colonnes_presentes = [col for col in colonnes_export if col in df_operations.columns]
            df_export = df_operations[colonnes_presentes]
            
            # Renommer les colonnes en français
            renommage = {
                'nom': 'Nom de l\'opération',
                'type_operation': 'Type',
                'statut_principal': 'Statut',
                'responsable_aco': 'Responsable ACO',
                'commune': 'Commune',
                'pourcentage_avancement': 'Avancement (%)',
                'date_creation': 'Date de création',
                'score_risque': 'Score de risque',
                'rem_annuelle_prevue': 'REM annuelle prévue (€)',
                'budget_total': 'Budget total (€)',
                'nb_logements_total': 'Nb logements'
            }
            
            df_export = df_export.rename(columns=renommage)
            
            # Formatage des valeurs
            for col in df_export.columns:
                if 'REM' in col or 'Budget' in col:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}".replace(",", " ") if pd.notnull(x) and x != 0 else "0")
                elif 'Date' in col:
                    df_export[col] = df_export[col].apply(lambda x: format_date_fr(str(x)) if pd.notnull(x) else "")
                elif 'Avancement' in col or 'Score' in col:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:.1f}" if pd.notnull(x) else "0.0")
            
            # Écrire dans Excel
            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws1.append(r)
            
            # Mise en forme
            for col in ws1.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws1.column_dimensions[column].width = adjusted_width
            
            # En-tête en gras avec couleur
            for cell in ws1[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Feuille 2 : Statistiques détaillées
        ws2 = wb.create_sheet("Statistiques")
        
        if operations:
            # Titre
            ws2.append(['TABLEAU DE BORD SPIC 2.0'])
            ws2['A1'].font = Font(bold=True, size=16)
            ws2['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws2['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws2.append([''])  # Ligne vide
            
            # Statistiques par statut
            ws2.append(['Répartition par statut'])
            ws2['A3'].font = Font(bold=True)
            ws2.append(['Statut', 'Nombre', 'Pourcentage'])
            
            statuts = {}
            for op in operations:
                statut = op.get('statut_principal', 'Non défini')
                statuts[statut] = statuts.get(statut, 0) + 1
            
            total_ops = len(operations)
            for statut, count in statuts.items():
                pourcentage = (count / total_ops) * 100
                ws2.append([statut, count, f"{pourcentage:.1f}%"])
            
            ws2.append([''])  # Ligne vide
            
            # Statistiques par type
            ws2.append(['Répartition par type d\'opération'])
            ws2[f'A{ws2.max_row}'].font = Font(bold=True)
            ws2.append(['Type', 'Nombre', 'REM Totale (€)'])
            
            types_stats = {}
            for op in operations:
                type_op = op.get('type_operation', 'Non défini')
                if type_op not in types_stats:
                    types_stats[type_op] = {'count': 0, 'rem': 0}
                types_stats[type_op]['count'] += 1
                types_stats[type_op]['rem'] += op.get('rem_annuelle_prevue', 0)
            
            for type_op, stats in types_stats.items():
                rem_formate = f"{stats['rem']:,.0f}".replace(",", " ")
                ws2.append([type_op, stats['count'], rem_formate])
            
            ws2.append([''])  # Ligne vide
            
            # KPI globaux
            ws2.append(['INDICATEURS CLÉS'])
            ws2[f'A{ws2.max_row}'].font = Font(bold=True)
            
            # Calculs KPI
            avancement_moyen = sum(op.get('pourcentage_avancement', 0) for op in operations) / len(operations)
            rem_totale = sum(op.get('rem_annuelle_prevue', 0) for op in operations)
            budget_total = sum(op.get('budget_total', 0) for op in operations)
            ops_bloquees = len([op for op in operations if op.get('has_active_blocage', False)])
            
            kpis = [
                ['Avancement moyen', f"{avancement_moyen:.1f}%"],
                ['REM totale portfolio', f"{rem_totale:,.0f} €".replace(",", " ")],
                ['Budget total portfolio', f"{budget_total:,.0f} €".replace(",", " ")],
                ['Opérations bloquées', ops_bloquees],
                ['Date export', datetime.now().strftime('%d/%m/%Y %H:%M')]
            ]
            
            for kpi in kpis:
                ws2.append(kpi)
                
        # Feuille 3 : Phases si données disponibles
        if phases_data:
            ws3 = wb.create_sheet("Phases")
            ws3.append(['Opération', 'Phase Principale', 'Sous-phase', 'Ordre', 'Statut', 'Domaine', 'Responsable'])
            
            # En-tête en gras
            for cell in ws3[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for operation_id, phases in phases_data.items():
                operation_nom = "N/A"
                # Trouver le nom de l'opération
                for op in operations:
                    if op.get('id') == operation_id:
                        operation_nom = op.get('nom', 'N/A')
                        break
                
                for phase in phases:
                    statut = "Validée" if phase.get('est_validee', False) else "En cours"
                    if phase.get('blocage_actif', False):
                        statut = "Bloquée"
                    
                    ws3.append([
                        operation_nom,
                        phase.get('phase_principale', ''),
                        phase.get('sous_phase', ''),
                        phase.get('ordre_phase', 0),
                        statut,
                        phase.get('domaine', 'OPERATIONNEL'),
                        phase.get('responsable_principal', '')
                    ])
        
        # Sauvegarder dans BytesIO pour Streamlit
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"❌ Erreur export Excel : {e}")
        return BytesIO()

def validate_phase_data_streamlit(phase: Dict) -> List[str]:
    """Valide les données d'une phase et retourne les erreurs (version Streamlit)"""
    
    erreurs = []
    
    try:
        # Vérifications obligatoires
        if not phase.get('sous_phase', '').strip():
            erreurs.append("Le nom de la sous-phase est obligatoire")
        
        if not phase.get('phase_principale', '').strip():
            erreurs.append("La phase principale est obligatoire")
        
        # Vérifications des dates
        date_debut = phase.get('date_debut_prevue')
        date_fin = phase.get('date_fin_prevue')
        
        if date_debut and date_fin:
            try:
                debut = datetime.strptime(str(date_debut), '%Y-%m-%d').date()
                fin = datetime.strptime(str(date_fin), '%Y-%m-%d').date()
                
                if fin < debut:
                    erreurs.append("La date de fin ne peut pas être antérieure à la date de début")
                    
                if (fin - debut).days > 730:  # 2 ans
                    erreurs.append("La durée de la phase semble excessive (> 2 ans)")
                    
            except ValueError:
                erreurs.append("Format de date invalide")
        
        # Vérifications des durées
        duree_mini = phase.get('duree_mini_jours', 0)
        duree_maxi = phase.get('duree_maxi_jours', 0)
        
        if duree_mini and duree_maxi:
            if duree_mini > duree_maxi:
                erreurs.append("La durée minimale ne peut pas être supérieure à la durée maximale")
            
            if duree_maxi > 730:  # 2 ans
                erreurs.append("La durée maximale semble excessive (> 2 ans)")
            
            if duree_mini < 1:
                erreurs.append("La durée minimale doit être d'au moins 1 jour")
        
        # Vérification du domaine
        domaine = phase.get('domaine', '')
        if domaine and domaine not in config.DOMAINES_OPERATIONNELS:
            erreurs.append(f"Domaine '{domaine}' non reconnu")
        
        return erreurs
        
    except Exception as e:
        return [f"Erreur de validation : {str(e)}"]

def create_phase_summary_card(phase: Dict, show_details: bool = False) -> str:
    """Crée une carte HTML pour afficher une phase (pour Streamlit)"""
    
    try:
        # Déterminer l'état et la couleur
        if phase.get('blocage_actif', False):
            couleur_border = "#ef4444"
            couleur_bg = "#fef2f2"
            statut_icone = "🔴"
            statut_text = "BLOQUÉ"
        elif phase.get('est_validee', False):
            couleur_border = "#10b981"
            couleur_bg = "#f0fdf4"
            statut_icone = "✅"
            statut_text = "VALIDÉ"
        else:
            # Vérifier si en retard
            today = datetime.now().date()
            en_retard = False
            if phase.get('date_fin_prevue'):
                try:
                    date_fin = datetime.strptime(phase['date_fin_prevue'], '%Y-%m-%d').date()
                    en_retard = today > date_fin
                except:
                    pass
            
            if en_retard:
                couleur_border = "#ef4444"
                couleur_bg = "#fef2f2"
                statut_icone = "⏰"
                statut_text = "EN RETARD"
            else:
                couleur_border = "#f59e0b"
                couleur_bg = "#fffbeb"
                statut_icone = "⏳"
                statut_text = "EN COURS"
        
        # Icône selon le domaine
        domaine = phase.get('domaine', 'OPERATIONNEL')
        domaine_info = config.DOMAINES_OPERATIONNELS.get(domaine, {})
        icone_domaine = domaine_info.get('icone', '📌')
        
        # Construire la carte
        html = f"""
        <div style="border-left: 4px solid {couleur_border}; background: {couleur_bg}; padding: 15px; border-radius: 8px; margin: 10px 0; font-family: Arial, sans-serif;">
            <div style="display: flex; justify-content: between; align-items: center; margin-bottom: 10px;">
                <div style="display: flex; align-items: center; gap: 8px;">
                    <span style="font-size: 18px;">{icone_domaine}</span>
                    <strong style="color: #1f2937;">{phase.get('sous_phase', 'Phase')}</strong>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <span style="font-size: 16px;">{statut_icone}</span>
                    <span style="font-size: 12px; font-weight: bold; color: {couleur_border};">{statut_text}</span>
                </div>
            </div>
        """
        
        if show_details:
            html += f"""
            <div style="color: #6b7280; font-size: 12px; margin-bottom: 8px;">
                <strong>Phase:</strong> {phase.get('phase_principale', 'N/A')} | 
                <strong>Domaine:</strong> {domaine_info.get('nom', domaine)}
            </div>
            """
            
            if phase.get('responsable_principal'):
                html += f"""
                <div style="color: #6b7280; font-size: 12px; margin-bottom: 8px;">
                    <strong>Responsable:</strong> {phase['responsable_principal']}
                </div>
                """
            
            if phase.get('date_fin_prevue'):
                date_fin_fr = format_date_fr(phase['date_fin_prevue'])
                html += f"""
                <div style="color: #6b7280; font-size: 12px; margin-bottom: 8px;">
                    <strong>Échéance:</strong> {date_fin_fr}
                </div>
                """
            
            if phase.get('impact_rem', False):
                html += f"""
                <div style="color: #10b981; font-size: 12px; margin-bottom: 8px;">
                    💰 <strong>Impact REM:</strong> {phase.get('rem_impact_desc', 'Cette phase impacte la rentabilité')}
                </div>
                """
        
        html += "</div>"
        
        return html
        
    except Exception as e:
        st.error(f"❌ Erreur création carte phase : {e}")
        return f"<p>Erreur affichage phase: {str(e)}</p>"

# ============================================================================
# FONCTIONS DE TEST ET VALIDATION STREAMLIT
# ============================================================================

def validate_utils_functions_streamlit() -> bool:
    """Valide le bon fonctionnement des fonctions utilitaires pour Streamlit"""
    
    try:
        # Test formatage devise
        test_currency = format_currency(1234.56)
        assert "1 234,56 €" in test_currency
        
        # Test formatage date
        test_date = format_date_fr("2024-03-15")
        assert test_date == "15/03/2024"
        
        # Test calcul avancement
        phases_test = [
            {'est_validee': True},
            {'est_validee': True},
            {'est_validee': False},
            {'est_validee': False}
        ]
        test_progress = calculate_progress(phases_test)
        assert test_progress == 50.0
        
        # Test icônes phases
        phase_test = {'sous_phase': 'travaux fondations', 'domaine': 'OPERATIONNEL'}
        test_icon = get_phase_icon(phase_test)
        assert test_icon in ['🏗️', '🚧', '📌']  # Icônes possibles
        
        return True
        
    except Exception as e:
        st.error(f"❌ Erreur validation utils : {e}")
        return False

def get_demo_data_for_streamlit() -> Dict:
    """Retourne des données de démonstration pour tester l'interface Streamlit"""
    
    return {
        'operations': [
            {
                'id': 1,
                'nom': '44 COUR CHARNEAU',
                'type_operation': 'OPP',
                'responsable_aco': 'Merezia CALVADOS',
                'statut_principal': '🚧 En travaux',
                'pourcentage_avancement': 57.8,
                'score_risque': 35.2,
                'rem_annuelle_prevue': 117600,
                'commune': 'LES ABYMES',
                'nb_logements_total': 18,
                'budget_total': 2800000,
                'has_active_blocage': False,
                'date_creation': '2024-01-15'
            }
        ],
        'kpi_data': {
            'total_operations': 1,
            'avancement_moyen': 57.8,
            'operations_bloquees': 0,
            'alertes_actives': 2,
            'repartition_statuts': {'🚧 En travaux': 1},
            'repartition_types': {'OPP': 1},
            'rem_totale_portfolio': 117600
        }
    }

# ============================================================================
# CACHE MANAGEMENT POUR STREAMLIT
# ============================================================================

def clear_streamlit_cache():
    """Vide le cache Streamlit pour forcer la mise à jour des données"""
    try:
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"❌ Erreur vidage cache : {e}")
        return False

def get_cache_info() -> Dict:
    """Retourne des informations sur le cache Streamlit"""
    try:
        return {
            'cache_enabled': True,
            'cache_ttl_default': 300,  # 5 minutes
            'cache_functions': [
                'calculate_progress',
                'generate_timeline',
                'generate_kpi_charts_streamlit',
                'generate_operation_summary_cached',
                'generate_aco_performance_summary_streamlit'
            ]
        }
    except Exception as e:
        st.error(f"❌ Erreur info cache : {e}")
        return {}

# Test du module si exécuté directement
if __name__ == "__main__":
    print("🧪 Test du module utils SPIC 2.0 Streamlit")
    
    # Validation des fonctions
    validation_ok = validate_utils_functions_streamlit()
    
    if validation_ok:
        print("✅ Module utils SPIC 2.0 Streamlit opérationnel")
    else:
        print("❌ Erreurs détectées dans le module utils")
    
    # Test de données de démonstration
    demo_data = get_demo_data_for_streamlit()
    print(f"📊 Données de demo préparées : {len(demo_data['operations'])} opération(s)")
    
    # Info cache
    cache_info = get_cache_info()
    print(f"💾 Fonctions en cache : {len(cache_info.get('cache_functions', []))}")
    
    print("📈 Module prêt pour intégration avec app.py Streamlit")